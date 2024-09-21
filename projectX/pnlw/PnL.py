import time
import tls_client
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook  # For Excel output
from openpyxl.styles import Font

# TLS session setup
session = tls_client.Session(
    client_identifier="chrome112",
    random_tls_extension_order=True
)

# Headers for making requests
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.5615.137 Safari/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Referer': 'https://gmgn.ai/',
    'TE': 'Trailers'
}

# Function to get wallet statistics
def get_wallet_stats(wallet: str):
    walletEndpoint = f"https://gmgn.ai/defi/quotation/v1/smartmoney/sol/walletNew/{wallet}?period=7d"
    response = session.get(walletEndpoint, headers=headers)

    if response.status_code == 200:
        data = response.json().get('data', {})
        if data:
            try:
                winrate_value = float(data.get('winrate') or 0) * 100  # Convert to percentage
                roi_value = float(data.get('total_profit_pnl') or 0) * 100  # Convert to percentage
                sol_value = float(data.get('sol_balance') or 0.0)
                buy = int(data.get('buy_7d') or 0)
                sell = int(data.get('sell_7d') or 0)
                txns = buy + sell
                token_avg_cost = float(data.get('token_avg_cost') or 0.0)
                pnl_lt_2x_num = int(data.get('pnl_lt_2x_num') or 0)
                pnl_2x_5x_num = int(data.get('pnl_2x_5x_num') or 0)
                pnl_gt_5x_num = int(data.get('pnl_gt_5x_num') or 0)
                last_active_timestamp = data.get('last_active_timestamp', 'N/A')

                return (wallet, winrate_value, roi_value, sol_value, txns, token_avg_cost, pnl_lt_2x_num, pnl_2x_5x_num, pnl_gt_5x_num, last_active_timestamp)
            except (ValueError, TypeError):
                return (wallet, 0.0, 0.0, 0.0, 0, 0.0, 0, 0, 0, 'N/A')
    return (wallet, 0.0, 0.0, 0.0, 0, 0.0, 0, 0, 0, 'N/A')

# Load wallet addresses from a file called wallets.txt
with open('wallets.txt', 'r') as fp:
    wallet_addresses = fp.read().splitlines()
    print(f"[✅] Loaded {len(wallet_addresses)} wallet addresses")

# Process wallets concurrently
threads = 15
print(f"[🤖] Set threads to {threads}")
print("[🔍] Fetching wallet statistics...")

startTime = time.time()

wallet_data = []

# Fetch wallet statistics in parallel
with ThreadPoolExecutor(max_workers=threads) as executor:
    futures = {executor.submit(get_wallet_stats, wallet): wallet for wallet in wallet_addresses}
    for future in as_completed(futures):
        wallet_stats = future.result()
        wallet_data.append(wallet_stats)

endTime = time.time()
totalTime = endTime - startTime
print(f"[✅] Data fetched in {totalTime:.2f} seconds")

# Save wallet data to Excel
def save_to_excel(wallet_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Wallet Stats"

    # Write header row
    ws.append(["Wallet Address", "Win Rate (%)", "ROI (%)", "SOL Balance", "Transactions", "Token Avg Cost", 
               "PnL < 2x", "PnL 2x-5x", "PnL > 5x", "Last Active", "Cielo Link", "GMGN Link"])

    # Create a hyperlink font style
    hyperlink_font = Font(color="0000FF", underline="single")

    # Write data rows
    for wallet, winrate, roi, sol_balance, transactions, token_avg_cost, pnl_lt_2x, pnl_2x_5x, pnl_gt_5x, last_active in wallet_data:
        cielo_link = f"https://app.cielo.finance/profile/{wallet}/pnl/tokens?timeframe=7d"
        gmgn_link = f"https://gmgn.ai/sol/address/{wallet}"

        # Safely format the numerical values, ensuring they're cast as floats
        try:
            row_data = [
                wallet, f"{float(winrate):.2f}", f"{float(roi):.2f}", f"{float(sol_balance):.2f}", transactions,
                f"{float(token_avg_cost):.2f}", pnl_lt_2x, pnl_2x_5x, pnl_gt_5x, last_active
            ]
        except (ValueError, TypeError):
            # Fallback in case of any unexpected type errors
            row_data = [wallet, "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", "N/A", last_active]

        # Append the data before links
        ws.append(row_data)

        # Add hyperlinks for Cielo and GMGN links with clickable text
        ws.cell(row=ws.max_row, column=11).hyperlink = cielo_link
        ws.cell(row=ws.max_row, column=11).value = "Cielo Profile"
        ws.cell(row=ws.max_row, column=11).font = hyperlink_font

        ws.cell(row=ws.max_row, column=12).hyperlink = gmgn_link
        ws.cell(row=ws.max_row, column=12).value = "GMGN Profile"
        ws.cell(row=ws.max_row, column=12).font = hyperlink_font

    # Save the workbook
    wb.save('wallets_pnL.xlsx')
    print("[✅] Excel file saved as 'wallets_pnL.xlsx'")

# Save the data
save_to_excel(wallet_data)
