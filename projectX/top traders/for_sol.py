from pprint import pprint
import time
import tls_client
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
import tls_client.exceptions
from openpyxl import Workbook  # For Excel output
from openpyxl.styles import Font

# TLS session setup
session = tls_client.Session(
    client_identifier="chrome112",
    random_tls_extension_order=True
)

addressFrequency = defaultdict(int)
addressToContract = defaultdict(set)  # Use a set to avoid duplicates
totalTraders = 0
MAX_RETRIES = 5
RETRY_BACKOFF = 2  # in seconds

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.5615.137 Safari/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Referer': 'https://gmgn.ai/',
    'TE': 'Trailers'
}

# Function to fetch top traders for a contract address
def fetch_top_traders(contract_address, limit):
    traders = []
    page = 1
    fetched = 0
    while fetched < limit:
        url = f"https://gmgn.ai/defi/quotation/v1/tokens/top_traders/sol/{contract_address}?orderby=profit&direction=desc&page={page}"
        attempts = 0
        while attempts < MAX_RETRIES:
            try:
                response = session.get(url, headers=headers)
                page += 1
                try:
                    data = response.json().get('data', [])
                except ValueError:
                    print(f"Non-JSON response received for {contract_address}: {response.text}")
                    break
                if not data:
                    break
                traders.extend(data[:limit - fetched])
                fetched += len(data)
                time.sleep(0.5)
                break
            except tls_client.exceptions.TLSClientExeption as e:
                attempts += 1
                print(f"Error fetching data for {contract_address} (Attempt {attempts}/{MAX_RETRIES}): {e}")
                time.sleep(RETRY_BACKOFF * attempts)
                if attempts == MAX_RETRIES:
                    print(f"Failed to fetch data for {contract_address} after {MAX_RETRIES} attempts.")
                    break
        if page > 30:
            break
    return traders

with open('tokens.txt', 'r') as fp:
    contractAddresses = fp.read().splitlines()
    print(f"[✅] Loaded {len(contractAddresses)} contract addresses")

try:
    threads = int(input("[❓] Threads: "))
except Exception:
    threads = 15

print(f"[🤖] Set threads to {threads}")

try:
    trader_limit = int(input("[❓] How many top traders to scan for each contract address: "))
except Exception:
    trader_limit = 500  # Default to 500 if input fails

print(f"[🤖] Set trader limit to {trader_limit}")

try:
    min_frequency = int(input("[❓] Minimum frequency for repeated addresses (default is 1): "))
except Exception:
    min_frequency = 1

try:
    min_winrate = float(input("[❓] Minimum win rate (default is 40%): "))
except Exception:
    min_winrate = 40.0

try:
    max_winrate = float(input("[❓] Maximum win rate (default is 90%): "))
except Exception:
    max_winrate = 90.0

try:
    min_roi = float(input("[❓] Minimum ROI (default is 10%): "))
except Exception:
    min_roi = 10.0

try:
    max_roi = float(input("[❓] Maximum ROI (default is 500%): "))
except Exception:
    max_roi = 500.0

try:
    min_sol_bal = float(input("[❓] Minimum SOL Balance (default is 2): "))
except Exception:
    min_sol_bal = 2

try:
    max_sol_bal = float(input("[❓] Maximum SOL Balance (default is 10000): "))
except Exception:
    max_sol_bal = 10000

try:
    min_txns = float(input("[❓] Minimum Transactions (default is 50): "))
except Exception:
    min_txns = 50

try:
    max_txns = float(input("[❓] Maximum Transactions (default is 1000): "))
except Exception:
    max_txns = 1000

print(f"[🤖] Set minimum frequency to {min_frequency}")
print(f"[🤖] Set minimum win rate to {min_winrate}%")
print(f"[🤖] Set maximum win rate to {max_winrate}%")
print(f"[🤖] Set minimum ROI to {min_roi}%")
print(f"[🤖] Set maximum ROI to {max_roi}%")
print(f"[🤖] Set minimum SOL Balance to {min_sol_bal}")
print(f"[🤖] Set maximum SOL Balance to {max_sol_bal}")
print(f"[🤖] Set minimum Transactions to {min_txns}")
print(f"[🤖] Set maximum Transactions to {max_txns}")

print("[🔍] Scanning...")

startTime = time.time()

# Run the fetch for all contract addresses concurrently
with ThreadPoolExecutor(max_workers=threads) as executor:
    futures = {executor.submit(fetch_top_traders, contractAddress, trader_limit): contractAddress for contractAddress in contractAddresses}
    
    for future in as_completed(futures):
        contractAddress = futures[future]
        response = future.result()
        totalTraders += len(response)
        
        for trader in response:
            address = trader['address']
            if contractAddress not in addressToContract[address]:
                addressFrequency[address] += 1 
                addressToContract[address].add(contractAddress)

endTime = time.time()
totalTime = endTime - startTime

print("[✅] Scan Completed")

# Filter repeated addresses based on the minimum frequency
repeatedAddresses = [address for address, count in addressFrequency.items() if count >= min_frequency]

if not repeatedAddresses:
    print("NO WALLETS FOUND")               
else:
    # Function to get wallet statistics, including new fields (Tags, Twitter Username)
    def get_wallet_stats(wallet: str):
        walletEndpoint = f"https://gmgn.ai/defi/quotation/v1/smartmoney/sol/walletNew/{wallet}?period=7d"
        response = session.get(walletEndpoint, headers=headers)

        if response.status_code == 200:
            data = response.json().get('data', {})
            if data:
                winrate_value = data.get('winrate')
                winrate = (winrate_value * 100) if winrate_value is not None else 0.0

                roi_value = data.get('total_profit_pnl')
                roi = (roi_value * 100) if roi_value is not None else 0.0

                sol_value = float(data.get('sol_balance'))
                sol = sol_value if sol_value is not None else 0.0

                buy = data.get('buy_7d')
                sell = data.get('sell_7d')
                txns = (buy + sell) if buy is not None and sell is not None else 0.0

                # New fields
                token_avg_cost = data.get('token_avg_cost', 0.0)
                pnl_lt_2x_num = data.get('pnl_lt_2x_num', 0)
                pnl_2x_5x_num = data.get('pnl_2x_5x_num', 0)
                pnl_gt_5x_num = data.get('pnl_gt_5x_num', 0)

                # Tags and Twitter username fields
                tags = data.get('tags', [])
                twitter_username = data.get('twitter_username', 'None')

                # Convert tags list to a string, if not empty
                tags_str = ', '.join(tags) if isinstance(tags, list) and tags else 'None'

                return (winrate, roi, sol, txns, token_avg_cost, pnl_lt_2x_num, pnl_2x_5x_num, pnl_gt_5x_num, tags_str, twitter_username)
        return (0.0, 0.0, 0.0, 0.0, 0.0, 0, 0, 0, 'None', 'None')

    valid_repeated_addresses = []

    # Run wallet statistics fetch concurrently
    with ThreadPoolExecutor(max_workers=threads) as executor:
        futures = {executor.submit(get_wallet_stats, address): address for address in repeatedAddresses}
        for future in as_completed(futures):
            address = futures[future]
            stats = future.result()
            winrate, roi, sol, txns = stats[:4]
            if (min_winrate <= winrate <= max_winrate) and (min_roi <= roi <= max_roi) and (min_sol_bal <= sol <= max_sol_bal) and (min_txns <= txns <= max_txns):
                valid_repeated_addresses.append((address, *stats))

    # Function to save valid repeated addresses to Excel with clickable links and additional fields
    def save_to_excel(valid_repeated_addresses):
        wb = Workbook()
        ws = wb.active
        ws.title = "Repeated Addresses"

        # Write header row
        ws.append(["Wallet Address", "Frequency", "Contracts", "Win Rate (%)", "ROI (%)", "SOL Balance", "Transactions",
                   "Token Avg Cost", "PnL < 2x", "PnL 2x-5x", "PnL > 5x", "Tags", "Twitter Username", "Cielo Link", "GMGN Link"])

        # Create a hyperlink font style
        hyperlink_font = Font(color="0000FF", underline="single")

        # Write data rows
        for address, winrate, roi, sol, txns, token_avg_cost, pnl_lt_2x_num, pnl_2x_5x_num, pnl_gt_5x_num, tags, twitter_username in valid_repeated_addresses:
            contracts_str = ', '.join(addressToContract[address])
            count = addressFrequency[address]
            cielo_link = f"https://app.cielo.finance/profile/{address}/pnl/tokens?timeframe=7d"
            gmgn_link = f"https://gmgn.ai/sol/address/{address}"

            # Create the row data with new fields
            row_data = [
                address, count, contracts_str, f"{winrate:.2f}", f"{roi:.2f}", f"{sol:.2f}", f"{txns:.2f}",
                f"{token_avg_cost:.2f}", pnl_lt_2x_num, pnl_2x_5x_num, pnl_gt_5x_num, tags, twitter_username
            ]

            # Append the data before links
            ws.append(row_data)

            # Add hyperlinks for Cielo and GMGN links with clickable text
            ws.cell(row=ws.max_row, column=14).hyperlink = cielo_link
            ws.cell(row=ws.max_row, column=14).value = "Cielo Profile"
            ws.cell(row=ws.max_row, column=14).font = hyperlink_font

            ws.cell(row=ws.max_row, column=15).hyperlink = gmgn_link
            ws.cell(row=ws.max_row, column=15).value = "GMGN Profile"
            ws.cell(row=ws.max_row, column=15).font = hyperlink_font

        # Save the workbook
        wb.save("repeated_addresses.xlsx")
        print("[✅] Saved repeated addresses to repeated_addresses.xlsx")

    # Save to Excel with the new fields
    if valid_repeated_addresses:
        save_to_excel(valid_repeated_addresses)
    else:
        print("NO WALLETS MATCHED THE CRITERIA")
    
    print(f"\n\nTotal Execution Time: {totalTime:.2f} seconds")
